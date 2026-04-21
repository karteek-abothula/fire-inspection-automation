import os
import re
import math
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT_FOLDER = r"example_folder\PDFs\2025"
OUTPUT_FILE = r"example_folder\fire_inspections_output_final.xlsx"

# Add any Excel workbooks here that need to be merged into the same output.
# Update these paths on your machine.
EXCEL_INPUT_FILES = [
    r"example_folder\PDFs\2025\excel_report1.xlsx"
]

rows = []

# -------------------------
# General helpers
# -------------------------

def is_blank(value):
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, str) and value.strip() == "":
        return True
    return False

def clean_text(text):
    if is_blank(text):
        return ""

    text = str(text)
    text = re.sub(r"-{2,}", "", text)
    text = re.sub(r"\s*[-–—]+\s*$", "", text)
    return re.sub(r"\s+", " ", text).strip()

def remove_trailing_noise(text):
    if is_blank(text):
        return ""

    noise_patterns = [
        r"\bRecommendations\b.*",
        r"\bAdditional Comments\b.*",
        r"\bADDITIONAL EXPLANATION\b.*",
        r"\bInspection Report\b.*"
    ]

    cleaned = str(text)
    for pat in noise_patterns:
        cleaned = re.sub(pat, "", cleaned, flags=re.IGNORECASE | re.DOTALL)

    return clean_text(cleaned)

def format_date_mmddyyyy(value):
    """
    Convert Excel/pandas dates like 2025-12-18 00:00:00 to 12/18/2025.
    Keeps already-good strings like 09/22/2025 unchanged.
    Returns blank if value is empty.
    """
    if is_blank(value):
        return ""

    text = clean_text(value)

    # already mm/dd/yyyy
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", text):
        dt = pd.to_datetime(text, errors="coerce")
        return dt.strftime("%m/%d/%Y") if pd.notna(dt) else text

    dt = pd.to_datetime(value, errors="coerce")
    if pd.notna(dt):
        return dt.strftime("%m/%d/%Y")

    return text

def get_year_from_path(path):
    parts = os.path.normpath(path).split(os.sep)
    for part in parts:
        if re.fullmatch(r"\d{4}", part):
            return part

    file_name = os.path.basename(path)
    match = re.search(r"(20\d{2})", file_name)
    if match:
        return match.group(1)

    return ""

def is_compliant_file(file_name):
    return os.path.splitext(file_name)[0].strip().lower().endswith("cc")

def should_skip(root_path, file_name):
    root_lower = root_path.lower()

    if "compliant" in root_lower:
        return True

    if is_compliant_file(file_name):
        return True

    return False

# -------------------------
# PDF helpers
# -------------------------

def extract_building_number(text):
    match = re.search(r"Building No\.\:\s*([A-Za-z0-9\-]+)", text, re.IGNORECASE)
    return clean_text(match.group(1)) if match else ""

def extract_region_code(text):
    """
    Fixes cases like:
    Region: None Building No.: 005
    Should return: None
    """
    match = re.search(
        r"Region:\s*(.*?)(?=\s*Building No\.\:|$)",
        text,
        re.IGNORECASE
    )
    return clean_text(match.group(1)) if match else ""

def extract_field_from_line(label, text):
    """
    Multiline-safe line extractor for lines like:
    Facility: SUNY Adirondack Escort:
    Building: Eisenhart Hall Inspected by: ...
    """
    pattern = rf"^{re.escape(label)}[ \t]*(.+)$"
    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)

    if not match:
        return ""

    value = match.group(1)
    value = re.split(r"Inspected by:", value, flags=re.IGNORECASE)[0]
    value = re.split(r"Escort:", value, flags=re.IGNORECASE)[0]

    return clean_text(value)

def extract_inspection_date(text):
    match = re.search(r"Inspection Date:\s*([0-9/]+)", text, re.IGNORECASE)
    return format_date_mmddyyyy(match.group(1)) if match else ""

def split_into_building_blocks(full_text):
    pattern = re.compile(
        r"(Region:\s*.*?Building No\.\:\s*.*?\nFacility:\s*.*?\nBuilding:\s*.*?)(?=\nRegion:\s*.*?Building No\.\:|\Z)",
        re.IGNORECASE | re.DOTALL
    )
    return [m.group(1).strip() for m in pattern.finditer(full_text)]

def extract_violation_text(block_text):
    match = re.search(
        r"Violation:[ \t]*(.*?)(?=\n[ \t]*Hazard Cls\.:)",
        block_text,
        re.IGNORECASE | re.DOTALL
    )
    return clean_text(match.group(1)) if match else ""

def extract_field_from_block(block_text, label, next_labels):
    """
    Important fix:
    use [ \\t]* instead of \\s* after the label so a blank field stays blank
    instead of swallowing the next line label like 'Code Refer.:'
    """
    next_pattern = "|".join(re.escape(lbl) for lbl in next_labels)

    pattern = re.compile(
        rf"{re.escape(label)}[ \t]*(.*?)(?=\n[ \t]*(?:{next_pattern})|\Z)",
        re.IGNORECASE | re.DOTALL
    )

    match = pattern.search(block_text)
    return clean_text(match.group(1)) if match else ""

def parse_violation_blocks(building_block):
    pattern = re.compile(
        r"^\s*(\d+)\s+Violation:\s*(.*?)(?=^\s*\d+\s+Violation:|\Z)",
        re.IGNORECASE | re.DOTALL | re.MULTILINE
    )

    matches = pattern.findall(building_block)
    parsed = []

    stop_labels_common = [
        "Violation:",
        "Recurring Violation:",
        "Recommendations",
        "Additional Comments",
        "ADDITIONAL EXPLANATION",
        "Inspection Report",
        "Agency:",
        "Facility:",
        "Building:",
        "Region:"
    ]

    for serial_number, block_text in matches:
        violation_text = extract_violation_text("Violation: " + block_text)

        hazard = extract_field_from_block(
            block_text,
            "Hazard Cls.:",
            ["Code Refer.:", "Abate By:", "Location:", "Description:"] + stop_labels_common
        )

        code_ref = extract_field_from_block(
            block_text,
            "Code Refer.:",
            ["Abate By:", "Location:", "Description:"] + stop_labels_common
        )

        due_date = extract_field_from_block(
            block_text,
            "Abate By:",
            ["Location:", "Description:"] + stop_labels_common
        )

        location = extract_field_from_block(
            block_text,
            "Location:",
            ["Description:"] + stop_labels_common
        )

        description = extract_field_from_block(
            block_text,
            "Description:",
            stop_labels_common
        )

        hazard = remove_trailing_noise(hazard)
        code_ref = remove_trailing_noise(code_ref)
        due_date = format_date_mmddyyyy(remove_trailing_noise(due_date))
        location = remove_trailing_noise(location)
        description = remove_trailing_noise(description)

        final_violation_value = violation_text if violation_text else clean_text(serial_number)

        parsed.append({
            "Violation": final_violation_value,
            "Hazard Classification": hazard,
            "Code Reference": code_ref,
            "Due Date": due_date,
            "Location": location,
            "Description": description
        })

    return parsed

# -------------------------
# Excel workbook helpers
# -------------------------

def find_value_to_right(row, label_idx):
    for j in range(label_idx + 1, len(row)):
        if not is_blank(row[j]):
            return row[j]
    return ""

def normalize_excel_label(value):
    text = clean_text(value)
    text = re.sub(r"\s+", " ", text)
    return text.lower()

def parse_downstate_workbook(excel_path):
    print(f"\nPROCESSING EXCEL WORKBOOK: {excel_path}")

    try:
        sheets = pd.read_excel(excel_path, sheet_name=None, header=None, engine="openpyxl")
    except Exception as e:
        print(f"  ERROR opening workbook: {e}")
        return

    inspection_year = get_year_from_path(excel_path)

    for sheet_name, df_sheet in sheets.items():
        print(f"  SHEET: {sheet_name}")

        try:
            sheet_rows = df_sheet.where(pd.notna(df_sheet), None).values.tolist()

            region_code = ""
            original_facility = ""
            original_building = sheet_name
            building_number = ""
            inspection_date = ""

            # ---- metadata scan ----
            for row in sheet_rows:
                for idx, cell in enumerate(row):
                    label = normalize_excel_label(cell)

                    if label == "region:" and region_code == "":
                        region_code = clean_text(find_value_to_right(row, idx))

                    elif label == "facility:" and original_facility == "":
                        original_facility = clean_text(find_value_to_right(row, idx))

                    elif label == "building:" and clean_text(original_building) == clean_text(sheet_name):
                        value = clean_text(find_value_to_right(row, idx))
                        if value:
                            original_building = value

                    elif label == "building no.:" and building_number == "":
                        building_number = clean_text(find_value_to_right(row, idx))

                    elif label == "inspection date:" and inspection_date == "":
                        inspection_date = format_date_mmddyyyy(find_value_to_right(row, idx))

            # Mapping rule:
            # Facility in original Excel -> Campus in output
            # Building in original Excel -> Facility in output
            excel_campus = original_facility
            excel_facility = original_building

            # Replace literal nan strings if they ever appear
            if excel_campus.lower() == "nan":
                excel_campus = ""
            if excel_facility.lower() == "nan":
                excel_facility = ""
            if region_code.lower() == "nan":
                region_code = ""
            if building_number.lower() == "nan":
                building_number = ""

            current_violation = None
            parsed_rows = []

            # ---- violation rows scan ----
            for row in sheet_rows:
                labels_in_row = [normalize_excel_label(c) for c in row]

                if any("recommendations" == lbl for lbl in labels_in_row):
                    break

                # start of a new violation row
                violation_idx = None
                violation_label = ""
                for idx, lbl in enumerate(labels_in_row):
                    if lbl in ["violation:", "recurring violation:"]:
                        violation_idx = idx
                        violation_label = lbl
                        break

                if violation_idx is not None:
                    if current_violation:
                        parsed_rows.append(current_violation)

                    serial_number = clean_text(row[violation_idx - 1]) if violation_idx > 0 else ""
                    violation_text = clean_text(find_value_to_right(row, violation_idx))
                    final_violation_value = violation_text if violation_text else serial_number

                    current_violation = {
                        "Violation": final_violation_value,
                        "Hazard Classification": "",
                        "Code Reference": "",
                        "Due Date": "",
                        "Location": "",
                        "Description": ""
                    }
                    continue

                if current_violation is None:
                    continue

                for idx, lbl in enumerate(labels_in_row):
                    if lbl == "hazard cls.:":
                        current_violation["Hazard Classification"] = remove_trailing_noise(
                            clean_text(find_value_to_right(row, idx))
                        )
                    elif lbl == "code refer.:":
                        current_violation["Code Reference"] = remove_trailing_noise(
                            clean_text(find_value_to_right(row, idx))
                        )
                    elif lbl == "abate by:":
                        current_violation["Due Date"] = format_date_mmddyyyy(
                            find_value_to_right(row, idx)
                        )
                    elif lbl == "location:":
                        current_violation["Location"] = remove_trailing_noise(
                            clean_text(find_value_to_right(row, idx))
                        )
                    elif lbl == "description:":
                        current_violation["Description"] = remove_trailing_noise(
                            clean_text(find_value_to_right(row, idx))
                        )

            if current_violation:
                parsed_rows.append(current_violation)

            real_violation_rows = [
                v for v in parsed_rows
                if any([
                    v["Hazard Classification"],
                    v["Code Reference"],
                    v["Due Date"],
                    v["Description"]
                ])
            ]

            if not real_violation_rows:
                print(f"    SKIPPED SHEET: {sheet_name}")
                continue

            print(f"    BUILDING: {excel_facility} | ROWS: {len(real_violation_rows)}")

            for v in real_violation_rows:
                rows.append({
                    "Violation": v["Violation"],
                    "Description": v["Description"],
                    "Code Reference": v["Code Reference"],
                    "Hazard Classification": v["Hazard Classification"],
                    "Due Date": v["Due Date"],
                    "Inspection Date": inspection_date,
                    "Facility": excel_facility,
                    "Is this a dorm or a housing unit? Yes/ No": "",
                    "Building Number(BN)": building_number,
                    "Campus": excel_campus,
                    "Region Code (RC)": region_code,
                    "Year": inspection_year,
                    "Source PDF": os.path.basename(excel_path),
                    "Source Path": excel_path,
                    "Source Sheet": sheet_name
                })

        except Exception as e:
            print(f"    ERROR in sheet '{sheet_name}': {e}")

# -------------------------
# PDF processing
# -------------------------

for root, _, files in os.walk(ROOT_FOLDER):
    for file in files:
        if not file.lower().endswith(".pdf"):
            continue

        if should_skip(root, file):
            print(f"SKIPPED COMPLIANT: {os.path.join(root, file)}")
            continue

        pdf_path = os.path.join(root, file)
        print(f"\nPROCESSING PDF: {pdf_path}")

        try:
            with pdfplumber.open(pdf_path) as pdf:
                full_text = "\n".join((page.extract_text() or "") for page in pdf.pages)

            building_blocks = split_into_building_blocks(full_text)
            inspection_year = get_year_from_path(pdf_path)

            print(f"  FOUND BUILDINGS: {len(building_blocks)}")

            for block in building_blocks:
                region_code = extract_region_code(block)
                building_number = extract_building_number(block)
                pdf_facility = extract_field_from_line("Facility:", block)
                pdf_building = extract_field_from_line("Building:", block)
                inspection_date = extract_inspection_date(block)

                excel_facility = pdf_building
                excel_campus = pdf_facility

                violations = parse_violation_blocks(block)

                real_violations = [
                    v for v in violations
                    if any([
                        v["Hazard Classification"],
                        v["Code Reference"],
                        v["Due Date"],
                        v["Description"]
                    ])
                ]

                if not real_violations:
                    print(f"    SKIPPED BUILDING: {excel_facility}")
                    continue

                print(f"    BUILDING: {excel_facility} | ROWS: {len(real_violations)}")

                for v in real_violations:
                    rows.append({
                        "Violation": v["Violation"],
                        "Description": v["Description"],
                        "Code Reference": v["Code Reference"],
                        "Hazard Classification": v["Hazard Classification"],
                        "Due Date": v["Due Date"],
                        "Inspection Date": inspection_date,
                        "Facility": excel_facility,
                        "Is this a dorm or a housing unit? Yes/ No": "",
                        "Building Number(BN)": building_number,
                        "Campus": excel_campus,
                        "Region Code (RC)": region_code,
                        "Year": inspection_year,
                        "Source PDF": file,
                        "Source Path": pdf_path,
                        "Source Sheet": ""
                    })

        except Exception as e:
            print(f"  ERROR: {file} -> {e}")

# -------------------------
# Excel workbook processing
# -------------------------

for excel_file in EXCEL_INPUT_FILES:
    if not os.path.exists(excel_file):
        print(f"\nSKIPPED EXCEL WORKBOOK (not found): {excel_file}")
        continue

    parse_downstate_workbook(excel_file)

# -------------------------
# Save final output
# -------------------------

df = pd.DataFrame(rows)

# Ensure consistent column order
desired_columns = [
    "Violation",
    "Description",
    "Code Reference",
    "Hazard Classification",
    "Due Date",
    "Inspection Date",
    "Facility",
    "Is this a dorm or a housing unit? Yes/ No",
    "Building Number(BN)",
    "Campus",
    "Region Code (RC)",
    "Year",
    "Source PDF",
    "Source Path",
    "Source Sheet"
]

for col in desired_columns:
    if col not in df.columns:
        df[col] = ""

df = df[desired_columns]

# Final cleanup to remove literal nan strings
for col in df.columns:
    df[col] = df[col].apply(lambda x: "" if clean_text(x).lower() == "nan" else x)

df.to_excel(OUTPUT_FILE, index=False)

# Convert the output range into an Excel table
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

if ws.max_row >= 2 and ws.max_column >= 1:
    table_ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName="FireInspectionTable", ref=table_ref)

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

wb.save(OUTPUT_FILE)

print("\nDone!")
print("Rows extracted:", len(df))
print("Saved to:", OUTPUT_FILE)